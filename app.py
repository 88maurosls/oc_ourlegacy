import io
import re
import unicodedata
import zipfile
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUTPUT_COLUMNS = [
    "SKU",
    "Product Name",
    "Variant",
    "Size",
    "Quantity",
    "Unit Price",
    "RRP",
]


HEADER_ALIASES = {
    "sku": {
        "sku", "item code", "itemcode", "article code", "articlecode",
        "product code", "productcode", "style code", "stylecode",
    },
    "variant_sku": {
        "variant sku", "variantsku", "variant code", "variantcode",
        "colour sku", "color sku",
    },
    "product_name": {
        "product name", "productname", "product", "style name", "stylename",
        "model name", "modelname", "article name", "articlename",
    },
    "variant": {
        "variant", "variant name", "variantname", "colour", "color",
        "colour name", "color name", "colourway", "colorway",
        "fabric", "material", "description variant", "variant description",
    },
    "sc": {
        "sc", "size scale", "sizescale", "size range", "sizerange",
        "scale code", "scalecode", "size code", "sizecode",
        "scala taglie", "scal taglie", "scala",
    },
    "unit_price": {
        "unit price", "unitprice", "wholesale price", "wholesaleprice",
        "wholesale", "purchase price", "purchaseprice", "buying price",
        "buyingprice", "unit cost", "unitcost", "net price", "netprice",
    },
    "rrp": {
        "rrp", "retail price", "retailprice", "recommended retail price",
        "recommendedretailprice", "msrp", "list price", "listprice",
    },
    "total_quantity": {
        "q", "qty", "quantity", "total qty", "totalqty",
        "total quantity", "totalquantity",
    },
}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ").strip().casefold()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def compact_text(value: Any) -> str:
    return normalize_text(value).replace(" ", "")


COMPACT_ALIASES = {
    canonical: {compact_text(alias) for alias in aliases}
    for canonical, aliases in HEADER_ALIASES.items()
}


def excel_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_number(value: Any) -> float | int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        number = float(value)
        return int(number) if number.is_integer() else number

    text = str(value).strip()
    if not text:
        return None

    text = re.sub(r"[^0-9,\.\-]", "", text)
    if text in {"", "-", ".", ","}:
        return None

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        number = float(text)
    except ValueError:
        return None
    return int(number) if number.is_integer() else number


def match_header(value: Any) -> str | None:
    compact = compact_text(value)
    if not compact:
        return None

    for canonical, aliases in COMPACT_ALIASES.items():
        if compact in aliases:
            return canonical

    # Fuzzy matching soltanto per intestazioni abbastanza lunghe, per evitare
    # che celle come X, S o Q vengano interpretate erroneamente.
    if len(compact) >= 5:
        best_name = None
        best_score = 0.0
        for canonical, aliases in COMPACT_ALIASES.items():
            for alias in aliases:
                if len(alias) < 5:
                    continue
                score = SequenceMatcher(None, compact, alias).ratio()
                if score > best_score:
                    best_name = canonical
                    best_score = score
        if best_score >= 0.86:
            return best_name

    return None


def headers_in_row(worksheet, row_number: int) -> tuple[dict[str, int], list[str]]:
    found: dict[str, int] = {}
    visible: list[str] = []
    max_columns = min(worksheet.max_column, 120)

    for col in range(1, max_columns + 1):
        value = worksheet.cell(row_number, col).value
        if value is None or str(value).strip() == "":
            continue
        visible.append(excel_text(value))
        canonical = match_header(value)
        if canonical and canonical not in found:
            found[canonical] = col

    return found, visible


def infer_missing_columns(found: dict[str, int], visible_positions: dict[int, str]) -> dict[str, int]:
    result = dict(found)
    sc_col = result.get("sc")

    if sc_col:
        # Se il nome prodotto o la variante hanno intestazioni insolite, usa la
        # struttura tipica della tabella: le colonne descrittive sono prima di SC.
        descriptive_cols = [
            col for col, value in visible_positions.items()
            if col < sc_col and normalize_text(value) not in {"comment", "comments", "note", "notes"}
        ]
        sku_cols = {
            col for key, col in result.items() if key in {"sku", "variant_sku"}
        }
        candidates = [col for col in descriptive_cols if col not in sku_cols]

        if "product_name" not in result and candidates:
            result["product_name"] = candidates[0]
        if "variant" not in result:
            after_product = [
                col for col in candidates
                if col > result.get("product_name", 0)
            ]
            if after_product:
                result["variant"] = after_product[0]

    return result


def find_header_row_and_sheet(workbook):
    best_candidate = None

    for worksheet in workbook.worksheets:
        max_scan_rows = min(worksheet.max_row, 300)
        for row_number in range(1, max_scan_rows + 1):
            found, visible = headers_in_row(worksheet, row_number)
            visible_positions = {
                col: excel_text(worksheet.cell(row_number, col).value)
                for col in range(1, min(worksheet.max_column, 120) + 1)
                if worksheet.cell(row_number, col).value not in (None, "")
            }
            found = infer_missing_columns(found, visible_positions)

            score = sum(
                weight for key, weight in {
                    "sc": 5,
                    "unit_price": 5,
                    "rrp": 5,
                    "product_name": 3,
                    "variant": 3,
                    "sku": 1,
                    "variant_sku": 1,
                }.items() if key in found
            )

            candidate = (score, worksheet, row_number, found, visible)
            if best_candidate is None or score > best_candidate[0]:
                best_candidate = candidate

            required = {"sc", "unit_price", "rrp", "product_name", "variant"}
            if required.issubset(found):
                return worksheet, row_number, found

    if best_candidate:
        score, worksheet, row_number, found, visible = best_candidate
        detected = ", ".join(sorted(found.keys())) or "nessuna"
        row_preview = " | ".join(visible[:20]) or "riga vuota"
        raise ValueError(
            "Non riesco a identificare con certezza la riga delle intestazioni. "
            f"Miglior candidato: foglio '{worksheet.title}', riga {row_number}. "
            f"Campi riconosciuti: {detected}. Contenuto: {row_preview}"
        )

    raise ValueError("Il file non contiene fogli leggibili.")


def build_size_scales(worksheet, header_row: int, scale_column: int, quantity_columns: list[int]):
    scales: dict[str, dict[int, str]] = {}

    for row_number in range(1, header_row):
        raw_code = worksheet.cell(row_number, scale_column).value
        code = normalize_text(raw_code)
        if not code:
            continue

        sizes: dict[int, str] = {}
        for col in quantity_columns:
            size = excel_text(worksheet.cell(row_number, col).value)
            if size:
                sizes[col] = size

        if sizes:
            scales[code] = sizes

    if not scales:
        raise ValueError(
            "Non è stata rilevata alcuna scala taglie sopra la tabella, "
            f"nella colonna {get_column_letter(scale_column)}."
        )

    return scales


def convert_excel(file_bytes: bytes):
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=False, read_only=False)
    worksheet, header_row, positions = find_header_row_and_sheet(workbook)

    sku_column = positions.get("sku")
    variant_sku_column = positions.get("variant_sku")
    if sku_column is None and variant_sku_column is None:
        raise ValueError("Non trovo né la colonna SKU né la colonna Variant SKU.")

    product_name_column = positions["product_name"]
    variant_column = positions["variant"]
    scale_column = positions["sc"]
    unit_price_column = positions["unit_price"]
    rrp_column = positions["rrp"]
    total_quantity_column = positions.get("total_quantity")

    if unit_price_column <= scale_column + 1:
        raise ValueError(
            "Non trovo le colonne delle quantità tra SC e Unit Price."
        )

    quantity_columns = list(range(scale_column + 1, unit_price_column))
    size_scales = build_size_scales(
        worksheet, header_row, scale_column, quantity_columns
    )

    records: list[dict[str, Any]] = []
    warnings: list[str] = []
    rows_with_products = 0

    for row_number in range(header_row + 1, worksheet.max_row + 1):
        sku = worksheet.cell(row_number, sku_column).value if sku_column else None
        if (sku is None or str(sku).strip() == "") and variant_sku_column:
            sku = worksheet.cell(row_number, variant_sku_column).value

        product_name = worksheet.cell(row_number, product_name_column).value
        variant = worksheet.cell(row_number, variant_column).value
        scale_raw = worksheet.cell(row_number, scale_column).value

        if product_name is None or str(product_name).strip() == "":
            continue

        rows_with_products += 1
        scale_code = normalize_text(scale_raw)
        if not scale_code:
            warnings.append(f"Riga {row_number}: SC mancante, riga ignorata.")
            continue

        scale = size_scales.get(scale_code)
        if scale is None:
            warnings.append(
                f"Riga {row_number}: scala taglie SC '{excel_text(scale_raw)}' non trovata."
            )
            continue

        unit_price_raw = worksheet.cell(row_number, unit_price_column).value
        rrp_raw = worksheet.cell(row_number, rrp_column).value
        unit_price = parse_number(unit_price_raw)
        rrp = parse_number(rrp_raw)
        if unit_price is None:
            unit_price = excel_text(unit_price_raw)
        if rrp is None:
            rrp = excel_text(rrp_raw)

        row_quantity = 0.0

        for col in quantity_columns:
            quantity = parse_number(worksheet.cell(row_number, col).value)
            if quantity is None or quantity == 0:
                continue
            if quantity < 0:
                warnings.append(
                    f"Riga {row_number}: quantità negativa nella colonna "
                    f"{get_column_letter(col)}, ignorata."
                )
                continue

            size = scale.get(col)
            if not size:
                warnings.append(
                    f"Riga {row_number}: quantità presente nella colonna "
                    f"{get_column_letter(col)}, ma la scala SC "
                    f"'{excel_text(scale_raw)}' non contiene una taglia in quella posizione."
                )
                continue

            clean_quantity = int(quantity) if float(quantity).is_integer() else quantity
            records.append(
                {
                    "SKU": excel_text(sku),
                    "Product Name": excel_text(product_name),
                    "Variant": excel_text(variant),
                    "Size": size,
                    "Quantity": clean_quantity,
                    "Unit Price": unit_price,
                    "RRP": rrp,
                }
            )
            row_quantity += float(quantity)

        if total_quantity_column:
            expected_quantity = parse_number(
                worksheet.cell(row_number, total_quantity_column).value
            )
            if expected_quantity is not None and abs(
                row_quantity - float(expected_quantity)
            ) > 0.000001:
                warnings.append(
                    f"Riga {row_number}: totale taglie {row_quantity:g}, "
                    f"ma la colonna quantità indica {float(expected_quantity):g}."
                )

    if rows_with_products == 0:
        raise ValueError("Ho trovato le intestazioni, ma nessuna riga prodotto.")
    if not records:
        raise ValueError("Non è stata trovata alcuna quantità maggiore di zero da esportare.")

    detected_scales = [excel_text(code).upper() for code in size_scales.keys()]
    return records, warnings, worksheet.title, sorted(detected_scales), header_row


def validate_output_excel(output_bytes: bytes):
    with zipfile.ZipFile(io.BytesIO(output_bytes), "r") as archive:
        damaged = archive.testzip()
        if damaged:
            raise ValueError(f"Il file Excel generato è danneggiato: {damaged}")
        if any(name.startswith("xl/tables/") for name in archive.namelist()):
            raise ValueError("Controllo interno fallito: è presente una tabella Excel strutturata.")

    check_workbook = load_workbook(io.BytesIO(output_bytes), data_only=False)
    check_sheet = check_workbook.active
    headers = [check_sheet.cell(1, col).value for col in range(1, 8)]
    if headers != OUTPUT_COLUMNS:
        raise ValueError("Controllo interno fallito: intestazioni di output errate.")

    for row in range(2, check_sheet.max_row + 1):
        quantity_cell = check_sheet.cell(row, 5)
        if quantity_cell.value is not None and quantity_cell.number_format != "0":
            raise ValueError("Controllo interno fallito: formato Quantity non intero.")


def create_output_excel(records: list[dict[str, Any]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Linear"

    worksheet.append(OUTPUT_COLUMNS)
    for record in records:
        worksheet.append([record[column] for column in OUTPUT_COLUMNS])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:G{worksheet.max_row}"

    for row in range(2, worksheet.max_row + 1):
        worksheet.cell(row, 5).number_format = "0"
        worksheet.cell(row, 6).number_format = "#,##0.00"
        worksheet.cell(row, 7).number_format = "#,##0.00"

    widths = {
        "A": 18,
        "B": 24,
        "C": 42,
        "D": 12,
        "E": 12,
        "F": 14,
        "G": 14,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    output = io.BytesIO()
    workbook.save(output)
    output_bytes = output.getvalue()
    validate_output_excel(output_bytes)
    return output_bytes


def main():
    import streamlit as st

    st.set_page_config(page_title="Excel Size Matrix to Linear", page_icon="📊")
    st.title("Conversione ordine Excel in formato lineare")
    st.write(
        "Carica il file ordine. L'app individua la riga delle intestazioni, "
        "legge la scala taglie indicata nella colonna **SC** e crea una riga "
        "per ogni taglia con quantità maggiore di zero."
    )

    uploaded_file = st.file_uploader(
        "Carica un file .xlsx o .xlsm",
        type=["xlsx", "xlsm"],
    )

    if uploaded_file is not None:
        try:
            records, warnings, sheet_name, detected_scales, header_row = convert_excel(
                uploaded_file.getvalue()
            )
            output_bytes = create_output_excel(records)

            total_quantity = sum(float(record["Quantity"]) for record in records)
            col1, col2, col3 = st.columns(3)
            col1.metric("Righe esportate", len(records))
            col2.metric("Quantità totale", f"{total_quantity:g}")
            col3.metric("Foglio letto", sheet_name)

            st.caption(
                f"Intestazioni rilevate alla riga {header_row}. "
                "Scale taglie rilevate: " + ", ".join(detected_scales)
            )
            st.dataframe(records[:200], use_container_width=True, hide_index=True)

            source_name = Path(uploaded_file.name).stem
            st.download_button(
                label="Scarica Excel lineare",
                data=output_bytes,
                file_name=f"{source_name}_linear.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )

            if warnings:
                with st.expander(f"Avvisi di controllo ({len(warnings)})"):
                    for warning in warnings[:100]:
                        st.write("•", warning)
                    if len(warnings) > 100:
                        st.write(f"Altri {len(warnings) - 100} avvisi non mostrati.")

        except Exception as exc:
            st.error(f"Impossibile convertire il file: {exc}")


if __name__ == "__main__":
    main()
